# -*- coding: utf-8 -*-
"""
Excel 与 CSV 双向转换工具

用途：
- XLS/XLSX → CSV：将 Excel 转换为 CSV，便于在 Cursor 中查看和 AI 分析
- CSV → XLS：将 CSV 转换为 Excel（先 openpyxl 生成带批注的 .xlsx，再通过 Excel COM 转为 .xls）

流程说明：
1. Python（openpyxl）生成 .xlsx，表头含批注（第1-12列）
2. 调用 Excel COM 接口将 .xlsx 转为 .xls（.xls 格式不支持批注，批注会丢失，仅保留数据）

模板说明：
- 输出14列：1-12业务列含批注，M/N列保留但不写批注，O列已删除

使用方法：
1. 拖放（推荐）：将 .xlsx 文件拖到 快速转换Excel为CSV.bat 上
2. 命令行 XLSX→CSV：python 其他/文件转换/Excel转CSV工具.py <Excel文件路径> [输出csv路径]
3. 命令行 CSV→XLS：python 其他/文件转换/Excel转CSV工具.py --to-xls <CSV文件路径> [输出xls路径]
4. 命令行 XLSX→XLS：python 其他/文件转换/Excel转CSV工具.py --xlsx-to-xls <XLSX文件路径> [输出xls路径]
5. 批量：python 其他/文件转换/Excel转CSV工具.py 文件夹路径

依赖：openpyxl（xlsx）、xlrd（xls）、pywin32（COM 转 xls，仅 Windows）
"""
import csv
import os
import subprocess
import sys

openpyxl = xlrd = xlwt = xlutils_copy = win32com = None


def _ensure_package(package_name: str, import_name: str = None) -> bool:
    """缺失时自动 pip 安装，返回是否可用"""
    if import_name is None:
        import_name = package_name
    try:
        __import__(import_name)
        return True
    except ImportError:
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", package_name, "-q"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            __import__(import_name)
            return True
        except Exception:
            return False


def _ensure_deps():
    """确保 openpyxl、xlrd、pywin32 可用，缺失时自动安装"""
    global openpyxl, xlrd, xlwt, xlutils_copy, win32com
    if openpyxl is None:
        if _ensure_package("openpyxl"):
            import openpyxl as _ox
            openpyxl = _ox
    if xlrd is None:
        if _ensure_package("xlrd"):
            import xlrd as _xr
            import xlwt as _xw
            from xlutils.copy import copy as _xc
            xlrd, xlwt, xlutils_copy = _xr, _xw, _xc
    if win32com is None and sys.platform == "win32":
        if _ensure_package("pywin32", "win32com.client"):
            import win32com.client as _wc
            win32com = _wc


def _get_template_path():
    """获取项目亏损报表模板路径"""
    root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    return os.path.join(root, "testcases", "CSV", "项目亏损报表.xls")


def xlsx_to_csv(xlsx_path: str, csv_path: str = None, sheet_index: int = 0) -> str:
    """
    将 xlsx 文件转换为 csv
    """
    _ensure_deps()
    if openpyxl is None:
        raise ImportError("请先安装 openpyxl：pip install openpyxl")

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    if not xlsx_path.lower().endswith(('.xlsx', '.xlsm')):
        raise ValueError(f"仅支持 .xlsx 和 .xlsm 格式: {xlsx_path}")

    if csv_path is None:
        base = os.path.splitext(xlsx_path)[0]
        csv_path = f"{base}.csv"

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb.worksheets[sheet_index]

    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else str(cell) for cell in row])

    wb.close()
    return csv_path


def xls_to_csv(xls_path: str, csv_path: str = None, sheet_index: int = 0) -> str:
    """
    将 .xls（Excel 97-2003）文件转换为 CSV。
    使用 xlrd 读取，输出 UTF-8 BOM，便于 Excel 正确识别中文。
    """
    _ensure_deps()
    if xlrd is None:
        raise ImportError("请先安装 xlrd：pip install xlrd")

    if not os.path.exists(xls_path):
        raise FileNotFoundError(f"文件不存在: {xls_path}")

    if not xls_path.lower().endswith('.xls'):
        raise ValueError(f"仅支持 .xls 格式: {xls_path}")

    if csv_path is None:
        base = os.path.splitext(xls_path)[0]
        csv_path = f"{base}.csv"

    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(sheet_index)

    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            writer.writerow(["" if c is None else (str(c).strip() if isinstance(c, str) else str(c)) for c in row])

    return csv_path


# 表头批注映射（仅第1-12列，M/N列不写批注，O列已删除）
HEADER_COMMENT_MAP = {
    '用例责任人': 'caseUser',
    '计划完成时间': 'askEndDate',
    '用例关键字': '5c46da7ce0ee77405a14a36f',
    '用例标识': '5cf6139ee0ee776533ddd2a4',
    '前置条件': '60825163e0ee772a9fa67c24',
    '名称': '5a43282dba4014a42391fa59',
    '关联业务': '5c2ebb66900add501a14147c',
    '步骤名称': 'name',
    '步骤描述': 'desc',
    '输入': 'input',
    '期望输出': 'expectOutput',
    '备注': 'notes',
}


def csv_to_xls(csv_path: str, xls_path: str = None, template_path: str = None) -> str:
    """
    将 CSV 文件转换为 Excel（.xlsx 格式，含表头批注）

    基于 项目亏损报表.xls 模板生成。第1-12列表头含批注，M/N列（13-14）保留列和表头但不写批注，O列已删除，共14列。
    输出 .xlsx 格式以支持表头批注（.xls 格式不支持批注）。

    参数:
        csv_path: csv 文件路径
        xls_path: 输出路径，默认与 csv 同目录同名，扩展名为 .xlsx
        template_path: 模板路径，默认 testcases/CSV/项目亏损报表.xls

    返回:
        生成的 xlsx 文件路径
    """
    _ensure_deps()
    if openpyxl is None:
        raise ImportError("请安装: pip install openpyxl")
    if xlrd is None:
        raise ImportError("请安装: pip install xlrd")

    from openpyxl.comments import Comment

    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"文件不存在: {csv_path}")

    if not csv_path.lower().endswith('.csv'):
        raise ValueError(f"仅支持 .csv 格式: {csv_path}")

    template_path = template_path or _get_template_path()
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板不存在: {template_path}")

    if xls_path is None:
        base = os.path.splitext(csv_path)[0]
        xls_path = f"{base}.xlsx"

    # 读取 CSV 数据
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        csv_rows = list(reader)

    if not csv_rows:
        raise ValueError("CSV 文件为空")

    ncols_template = 14  # 共14列：1-12业务列+M/N列，O列已删除

    # 使用 openpyxl 创建 xlsx，第1-12列表头含批注，M/N列不写批注
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入表头（第1行），仅第1-12列添加批注，M/N列（13-14）不写批注
    header_row = list(csv_rows[0])[:ncols_template]
    while len(header_row) < ncols_template:
        header_row.append('')
    for c_idx, val in enumerate(header_row):
        cell_val = val.strip() if val else ''
        ws.cell(row=1, column=c_idx + 1, value=cell_val)
        # 仅第1-12列（c_idx 0-11）添加批注，M/N列不写批注
        if c_idx < 12 and cell_val and cell_val in HEADER_COMMENT_MAP:
            ws.cell(row=1, column=c_idx + 1).comment = Comment(
                HEADER_COMMENT_MAP[cell_val], "POM"
            )

    # 写入数据行（从第2行开始）
    for r_idx, row in enumerate(csv_rows[1:], start=2):
        row = list(row)[:ncols_template]
        while len(row) < ncols_template:
            row.append('')
        for c_idx, val in enumerate(row):
            ws.cell(row=r_idx, column=c_idx + 1, value=val if val else '')

    wb.save(xls_path)
    return xls_path


def xlsx_to_xls_com(xlsx_path: str, xls_path: str = None) -> str:
    """
    通过 Excel COM 接口将 .xlsx 转为 .xls（Excel 97-2003 格式）

    注意：.xls 格式不支持批注，转换后批注会丢失，仅保留单元格数据。

    参数:
        xlsx_path: 源 xlsx 文件路径
        xls_path: 输出 xls 路径，默认与 xlsx 同目录同名，扩展名为 .xls

    返回:
        生成的 xls 文件路径
    """
    _ensure_deps()
    if win32com is None:
        raise ImportError("请安装 pywin32：pip install pywin32（仅 Windows，需已安装 Excel）")

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    if xls_path is None:
        base = os.path.splitext(xlsx_path)[0]
        xls_path = f"{base}.xls"

    xlsx_path = os.path.abspath(xlsx_path)
    xls_path = os.path.abspath(xls_path)

    xl = win32com.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(xlsx_path)
        wb.SaveAs(xls_path, FileFormat=56)  # 56 = xlExcel8 (.xls)
        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()

    return xls_path


def csv_to_xlsx(csv_path: str, xlsx_path: str = None) -> str:
    """
    将 CSV 转换为 xlsx（保留用于兼容，推荐使用 csv_to_xls）
    """
    _ensure_deps()
    if openpyxl is None:
        raise ImportError("请先安装 openpyxl：pip install openpyxl")
    from openpyxl.comments import Comment

    # 与 csv_to_xls 一致：仅第1-12列添加批注，M/N列不写批注，共14列
    _comment_map = {
        '用例责任人': 'caseUser',
        '计划完成时间': 'askEndDate',
        '用例关键字': '5c46da7ce0ee77405a14a36f',
        '用例标识': '5cf6139ee0ee776533ddd2a4',
        '前置条件': '60825163e0ee772a9fa67c24',
        '名称': '5a43282dba4014a42391fa59',
        '关联业务': '5c2ebb66900add501a14147c',
        '步骤名称': 'name',
        '步骤描述': 'desc',
        '输入': 'input',
        '期望输出': 'expectOutput',
        '备注': 'notes',
    }

    if xlsx_path is None:
        base = os.path.splitext(csv_path)[0]
        xlsx_path = f"{base}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row[:14])  # 只写入前14列

    if ws.max_row >= 1:
        for col_idx, cell in enumerate(ws[1], start=1):
            if col_idx <= 12 and cell.value and str(cell.value).strip() in _comment_map:
                cell.comment = Comment(_comment_map[str(cell.value).strip()], "POM")

    wb.save(xlsx_path)
    return xlsx_path


def convert_file(path: str, output_dir: str = None) -> str:
    """转换单个文件（支持 .xlsx/.xlsm/.xls）"""
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        base = os.path.basename(path)
        csv_path = os.path.join(output_dir, os.path.splitext(base)[0] + ".csv")
    else:
        csv_path = None
    if path.lower().endswith('.xls'):
        return xls_to_csv(path, csv_path)
    return xlsx_to_csv(path, csv_path)


def convert_folder(folder_path: str, output_dir: str = None) -> list:
    """批量转换文件夹内的 xlsx/xls 文件"""
    if output_dir is None:
        output_dir = folder_path
    results = []
    for name in os.listdir(folder_path):
        if name.lower().endswith(('.xlsx', '.xlsm', '.xls')):
            path = os.path.join(folder_path, name)
            try:
                out = convert_file(path, output_dir)
                results.append(out)
                print(f"✓ {name} -> {os.path.basename(out)}")
            except Exception as e:
                print(f"✗ {name}: {e}")
    return results


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\n示例：")
        print("  python 其他/文件转换/Excel转CSV工具.py 测试用例.xlsx")
        print("  python 其他/文件转换/Excel转CSV工具.py --to-xls 测试用例.csv  # CSV→xlsx(带批注)→xls(COM)")
        print("  python 其他/文件转换/Excel转CSV工具.py --xlsx-to-xls 测试用例.xlsx  # 仅 xlsx→xls(COM)")
        print("  python 其他/文件转换/Excel转CSV工具.py testcases/CSV/  # 批量转换")
        sys.exit(0)

    # CSV → XLS 模式（先 openpyxl 生成带批注的 xlsx，再 Excel COM 转为 xls）
    if sys.argv[1] == '--to-xls':
        if len(sys.argv) < 3:
            print("用法: python Excel转CSV工具.py --to-xls <CSV文件路径> [输出xls路径]")
            sys.exit(1)
        path = sys.argv[2].strip('"').strip("'")
        xls_out = sys.argv[3].strip('"').strip("'") if len(sys.argv) > 3 else None
        if os.path.isfile(path):
            try:
                # 1. Python 生成带批注的 .xlsx
                base = os.path.splitext(path)[0]
                xlsx_path = f"{base}.xlsx"
                csv_to_xls(path, xlsx_path)
                print(f"✓ 已生成带批注的 .xlsx: {xlsx_path}")
                # 2. Excel COM 转为 .xls
                out = xlsx_to_xls_com(xlsx_path, xls_out)
                print(f"✓ 已通过 Excel COM 转为 .xls: {out}")
            except Exception as e:
                print(f"✗ 转换失败: {e}")
                sys.exit(1)
        else:
            print(f"✗ 文件不存在: {path}")
            sys.exit(1)
        return

    # XLSX → XLS 模式（仅 Excel COM 转换，需已安装 pywin32）
    if sys.argv[1] == '--xlsx-to-xls':
        if len(sys.argv) < 3:
            print("用法: python Excel转CSV工具.py --xlsx-to-xls <XLSX文件路径> [输出xls路径]")
            sys.exit(1)
        path = sys.argv[2].strip('"').strip("'")
        xls_out = sys.argv[3].strip('"').strip("'") if len(sys.argv) > 3 else None
        if os.path.isfile(path):
            try:
                out = xlsx_to_xls_com(path, xls_out)
                print(f"✓ 转换完成: {out}")
            except Exception as e:
                print(f"✗ 转换失败: {e}")
                sys.exit(1)
        else:
            print(f"✗ 文件不存在: {path}")
            sys.exit(1)
        return

    # 兼容旧的 --to-xlsx 参数
    if sys.argv[1] == '--to-xlsx':
        if len(sys.argv) < 3:
            print("用法: python Excel转CSV工具.py --to-xlsx <CSV文件路径> [输出xlsx路径]")
            sys.exit(1)
        path = sys.argv[2].strip('"').strip("'")
        xlsx_out = sys.argv[3].strip('"').strip("'") if len(sys.argv) > 3 else None
        if os.path.isfile(path):
            try:
                out = csv_to_xlsx(path, xlsx_out)
                print(f"✓ 转换完成: {out}")
            except Exception as e:
                print(f"✗ 转换失败: {e}")
                sys.exit(1)
        else:
            print(f"✗ 文件不存在: {path}")
            sys.exit(1)
        return

    path = sys.argv[1].strip('"').strip("'")
    csv_out = sys.argv[2].strip('"').strip("'") if len(sys.argv) > 2 else None

    if os.path.isfile(path):
        try:
            if path.lower().endswith('.xls'):
                out = xls_to_csv(path, csv_out)
            else:
                out = xlsx_to_csv(path, csv_out)
            print(f"✓ 转换完成: {out}")
        except Exception as e:
            print(f"✗ 转换失败: {e}")
            sys.exit(1)
    elif os.path.isdir(path):
        convert_folder(path, csv_out if csv_out and os.path.isdir(csv_out) else None)
    else:
        print(f"✗ 路径不存在: {path}")
        sys.exit(1)


if __name__ == "__main__":
    main()
